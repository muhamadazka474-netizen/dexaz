"""
Excel (.xlsx / .xlsm) import helpers.

Two-step flow, mirroring how the SQL Editor asks for confirmation before
anything destructive runs:

  1. analyze_workbook() — reads the file, guesses a column name + a
     PostgreSQL type for every column, and returns a small preview so the
     frontend can show the user what's about to happen and let them fix
     names/types before anything touches the database.
  2. extract_rows() — once the user confirms (optionally after editing
     column names/types/inclusion), re-reads the same file and returns the
     full row set for the columns that were kept, ready for a bulk insert.

Nothing here executes SQL — that's the adapter's job (see
DatabaseAdapter.import_rows).
"""
import io
import re
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import load_workbook


class ExcelImportError(Exception):
    """Raised for anything wrong with the uploaded file itself (bad format,
    empty sheet, out-of-range header row, ...) — always safe to show
    verbatim to the user."""


# Sampling caps: an import can be large, but the *preview* only ever needs
# to look at a bounded prefix of the sheet to guess types and show a few
# rows — the real row count is still reported via a full (cheap) scan.
MAX_PREVIEW_ROWS = 20
MAX_SAMPLE_ROWS_FOR_TYPE_INFERENCE = 500

_IDENT_INVALID = re.compile(r"[^a-z0-9_]+")
_IDENT_REPEAT_UNDERSCORE = re.compile(r"_+")


def _load(content: bytes):
    try:
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelImportError(f"Gagal membaca file Excel: {e}")


def _clean_identifier(raw: Any, idx: int) -> str:
    """Turns a header cell into a safe, lowercase Postgres identifier."""
    if raw is None or str(raw).strip() == "":
        return f"column_{idx + 1}"
    name = _IDENT_INVALID.sub("_", str(raw).strip().lower())
    name = _IDENT_REPEAT_UNDERSCORE.sub("_", name).strip("_")
    if not name:
        return f"column_{idx + 1}"
    if not re.match(r"^[a-z_]", name):
        name = f"col_{name}"
    return name


def _dedupe(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for n in names:
        if n not in seen:
            seen[n] = 0
            result.append(n)
        else:
            seen[n] += 1
            result.append(f"{n}_{seen[n]}")
    return result


def _infer_type(values: list[Any]) -> str:
    """Best-effort PostgreSQL type guess for a column, from sampled values."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "TEXT"
    if all(isinstance(v, bool) for v in non_null):
        return "BOOLEAN"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_null):
        return "BIGINT"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_null):
        return "DOUBLE PRECISION"
    if all(isinstance(v, datetime) for v in non_null):
        return "TIMESTAMP"
    if all(isinstance(v, date) for v in non_null):
        return "DATE"
    return "TEXT"


def _jsonify(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _resolve_sheet(wb, sheet_name: Optional[str]) -> str:
    sheets = wb.sheetnames
    if not sheets:
        raise ExcelImportError("File Excel tidak berisi sheet apa pun")
    if sheet_name and sheet_name in sheets:
        return sheet_name
    return sheets[0]


def list_sheets(content: bytes) -> list[str]:
    wb = _load(content)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def analyze_workbook(content: bytes, sheet_name: Optional[str] = None, header_row: int = 1) -> dict:
    """Reads the sheet, guesses column names/types, and returns a preview.
    Does not touch the database."""
    if header_row < 1:
        raise ExcelImportError("Baris header harus 1 atau lebih")

    wb = _load(content)
    try:
        sheets = wb.sheetnames
        active_sheet = _resolve_sheet(wb, sheet_name)
        ws = wb[active_sheet]

        sampled_rows: list[tuple] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sampled_rows.append(row)
            if i >= header_row + MAX_SAMPLE_ROWS_FOR_TYPE_INFERENCE:
                break

        if len(sampled_rows) < header_row:
            raise ExcelImportError(
                f"Sheet '{active_sheet}' tidak memiliki baris ke-{header_row} untuk dijadikan header"
            )

        header = sampled_rows[header_row - 1]
        data_rows = sampled_rows[header_row:]

        raw_names = [str(h) if h is not None else "" for h in header]
        cleaned = _dedupe([_clean_identifier(h, idx) for idx, h in enumerate(header)])

        columns = []
        for idx in range(len(header)):
            col_values = [r[idx] if idx < len(r) else None for r in data_rows]
            if header[idx] is None and all(v is None for v in col_values):
                continue  # fully empty column — most likely a stray blank cell in the header row
            columns.append(
                {
                    "index": idx,
                    "source_name": raw_names[idx] or f"Column {idx + 1}",
                    "target_name": cleaned[idx],
                    "type": _infer_type(col_values),
                    "include": True,
                }
            )

        preview_rows = [
            [_jsonify(r[c["index"]]) if c["index"] < len(r) else None for c in columns]
            for r in data_rows[:MAX_PREVIEW_ROWS]
        ]

        # A full (but cheap — just iterating, no cell parsing) scan for the
        # real row count, since the sample above is capped.
        total_data_rows = sum(1 for _ in ws.iter_rows(min_row=header_row + 1, values_only=True))

        return {
            "sheets": sheets,
            "active_sheet": active_sheet,
            "header_row": header_row,
            "columns": columns,
            "preview_rows": preview_rows,
            "row_count": total_data_rows,
        }
    finally:
        wb.close()


def extract_rows(
    content: bytes, sheet_name: Optional[str], header_row: int, columns: list[dict]
) -> list[list[Any]]:
    """Reads every data row for the given columns (as returned/edited from
    analyze_workbook), converting cells to plain Python values ready for a
    parameterized bulk insert. Rows that are entirely blank are skipped."""
    wb = _load(content)
    try:
        active_sheet = _resolve_sheet(wb, sheet_name)
        ws = wb[active_sheet]
        indices = [c["index"] for c in columns]

        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row:
                continue
            values = [row[idx] if idx < len(row) else None for idx in indices]
            if all(v is None for v in values):
                continue
            rows.append(values)
        return rows
    finally:
        wb.close()
