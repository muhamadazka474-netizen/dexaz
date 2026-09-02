"""
Format generators for exported data (spec section 31-32).

Each function takes columns + rows (already fetched from the adapter) and
returns raw bytes ready to stream back as a file download. Excel exports
include a second "Query Information" sheet documenting what was exported
and when, per spec.
"""
import csv
import io
import json
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def to_csv_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: ("" if row.get(c) is None else row.get(c)) for c in columns})
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 correctly


def to_json_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    return json.dumps(rows, default=str, indent=2, ensure_ascii=False).encode("utf-8")


def to_excel_bytes(
    columns: list[str],
    rows: list[dict[str, Any]],
    meta: dict[str, Any],
) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Query Result"
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(c) if row.get(c) is not None else "" for c in columns])
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)

    info_ws = wb.create_sheet("Query Information")
    info_ws.append(["Field", "Value"])
    for cell in info_ws[1]:
        cell.font = Font(bold=True)
    info_rows = [
        ("Database", meta.get("database", "")),
        ("Source", meta.get("source", "")),
        ("Executed At", meta.get("executed_at", datetime.now(timezone.utc).isoformat())),
        ("Row Count", len(rows)),
        ("SQL Query", meta.get("sql", "")),
    ]
    for label, value in info_rows:
        info_ws.append([label, value])
    info_ws.column_dimensions["A"].width = 16
    info_ws.column_dimensions["B"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


CONTENT_TYPES = {
    "csv": "text/csv",
    "json": "application/json",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
EXTENSIONS = {"csv": "csv", "json": "json", "excel": "xlsx"}


def generate(fmt: str, columns: list[str], rows: list[dict[str, Any]], meta: dict[str, Any]) -> bytes:
    if fmt == "csv":
        return to_csv_bytes(columns, rows)
    if fmt == "json":
        return to_json_bytes(columns, rows)
    if fmt == "excel":
        return to_excel_bytes(columns, rows, meta)
    raise ValueError(f"Unsupported export format: {fmt}")
